import openpyxl as OPX
import sys
import getopt
import os.path
import shutil
from os.path import expanduser
import pymysql
from database_functions import get_db_instance

col_name = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P']

def main(argv):
    # extract data from excel file
    tables,sheet_names = get_data(argv[0])
    # insert data into db
    conn = get_db_instance()
    populate_db(conn, tables, sheet_names)

def get_data(xl_file):
    # extract data from excel sheets into objects for tables
    wb = OPX.load_workbook(filename=xl_file)
    sheet_names = wb.get_sheet_names()
    print(sheet_names)

    table = {}
    for sheet in sheet_names:
        data = []
        ws = wb.get_sheet_by_name(sheet)
        iter_range = get_range(ws)
        for row in ws.iter_rows(iter_range):
            entry = {}
            for i,cell in enumerate(row):
                key = ws[col_name[i] + '1'].value
                entry[key] = cell.value
            data.append(entry)
        table[sheet] = data
    return(table, sheet_names)

def populate_db(conn, tables, sheet_names):
    try:
        with conn.cursor() as cur:
            for sheet in sheet_names:
                cur.execute('delete from '+str(sheet).lower() + ';')
                for record in tables[sheet]:
                    placeholders = ', '.join(['%s'] * len(record))
                    columns = ', '.join(record.keys())
                    sql = 'insert into %s (%s) Values (%s)'\
                        % (str(sheet).lower(), columns, placeholders)
                    cur.execute(sql, list(record.values()))
            conn.commit()
    finally:
        conn.close()

def get_range(ws):
    # Define appropriate range for iteration
    start_row = 2
    end_row = ws.max_row
    start_col = col_name[0]
    end_col = col_name[ws.max_column - 1]
    iter_range = "{start_col}{start_row}:{end_col}{end_row}".format(start_col=start_col,
        end_col=end_col, start_row=start_row, end_row=end_row)
    return(iter_range)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("add excel file as input")
        print("e.g. >> python populate.py [EXCEL FILE]")
    else:
        main(sys.argv[1:])
